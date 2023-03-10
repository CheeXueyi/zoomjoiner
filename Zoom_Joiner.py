import webbrowser
from datetime import datetime
from openpyxl import load_workbook
import time
import pynput
import subprocess
from win32api import GetSystemMetrics



#opening excel file
data = load_workbook(filename = "input.xlsx")
sheet = data.active



#function definition
def process_exists(process_name):
    call = 'TASKLIST', '/FI', 'imagename eq %s' % process_name
    # use buildin check_output right away
    output = subprocess.check_output(call).decode()
    # check in last line for process name
    last_line = output.strip().split('\r\n')[-1]
    # because Fail message could be translated
    return last_line.lower().startswith(process_name.lower())



#class definition
class zoomMeetingDetails:
    def __init__(self, link, password, day, jointime, leavetime):
        self.link = link
        self.password = password
        self.day = day
        self.jointime = jointime
        self.leavetime = leavetime



#global variables
classes = []
keyboard = pynput.keyboard.Controller()
mouse = pynput.mouse.Controller()
screenWidth = GetSystemMetrics(0)
screenHeight = GetSystemMetrics(1)
numOfSeconds = sheet['B1'].value


#placing class details in objects in list "classes"
for value in sheet.iter_rows(min_row=4, max_row=500, min_col=1, max_col=5, values_only=True):
    if value[0] == None:
        break
    link = value[0]
    password = value[1]
    day = value[2].upper()
    jointime = value[3]
    leavetime = value[4]
    classes.append(zoomMeetingDetails(link, password, day, jointime, leavetime))
    print(link)



#mainloop
print("classjoiner is now running")
while True:
    currentDay = datetime.now().strftime("%a").upper()
    currentHour = int(datetime.now().strftime("%H"))
    currentMinute = int(datetime.now().strftime("%M"))
    
    for i in classes:
        if i.jointime.hour == currentHour and i.jointime.minute == currentMinute and i.day == currentDay:
            subprocess.call(["taskkill", "/f", "/im", "ZOOM.EXE"])
            webbrowser.open(i.link)
            index = 0
            while (process_exists("zoom.exe") == False):
                if index == 0:
                    print("opening zoom")
                    index += 1
            time.sleep(numOfSeconds)
            mouse.position = (screenWidth/2, screenHeight/2)
            mouse.click(pynput.mouse.Button.left)
            keyboard.press(pynput.keyboard.Key.enter)
            keyboard.release(pynput.keyboard.Key.enter)
            time.sleep(numOfSeconds)
            mouse.position = (screenWidth/2, screenHeight/2)
            mouse.click(pynput.mouse.Button.left)
            keyboard.type(str(i.password))
            keyboard.press(pynput.keyboard.Key.enter)
            keyboard.release(pynput.keyboard.Key.enter)
            time.sleep(60)
        elif i.day == currentDay and i.leavetime.hour == currentHour and i.leavetime.minute == currentMinute:
            subprocess.call(["taskkill", "/f", "/im", "ZOOM.EXE"])
    time.sleep(30)